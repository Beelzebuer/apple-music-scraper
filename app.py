"""
Apple Music 专辑爬虫 Web 应用
Flask 后端 + 苹果风格前端
"""

import io
import json
import os
from datetime import datetime

from flask import Flask, jsonify, request, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from scraper import AppleMusicScraper

app = Flask(__name__)
scraper = AppleMusicScraper()


@app.route('/')
def index():
    """主页面"""
    return render_template('index.html')


@app.route('/api/scrape', methods=['POST'])
def api_scrape():
    """单条抓取 API"""
    data = request.get_json()
    url = data.get('url', '').strip()

    if not url:
        return jsonify({'success': False, 'error': 'URL 不能为空'}), 400

    try:
        result = scraper.scrape(url)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/scrape-batch', methods=['POST'])
def api_scrape_batch():
    """批量抓取 API"""
    data = request.get_json()
    urls = data.get('urls', [])

    # 过滤空行
    urls = [url.strip() for url in urls if url.strip()]

    if not urls:
        return jsonify({'success': False, 'error': '请提供至少一个有效的 URL'}), 400

    if len(urls) > 50:
        return jsonify({'success': False, 'error': '单次最多支持 50 个链接'}), 400

    results = scraper.scrape_batch(urls)

    success_count = sum(1 for r in results if r['success'])
    fail_count = len(results) - success_count

    return jsonify({
        'success': True,
        'results': results,
        'total': len(results),
        'success_count': success_count,
        'fail_count': fail_count
    })


@app.route('/api/download/json', methods=['POST'])
def download_json():
    """下载 JSON 文件"""
    data = request.get_json()
    albums = data.get('albums', [])

    if not albums:
        return jsonify({'success': False, 'error': '没有数据可下载'}), 400

    # 创建内存文件
    output = io.BytesIO()
    output.write(json.dumps(albums, ensure_ascii=False, indent=2).encode('utf-8'))
    output.seek(0)

    filename = f"apple_music_albums_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    return send_file(
        output,
        mimetype='application/json',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/download/excel', methods=['POST'])
def download_excel():
    """下载 Excel 文件"""
    data = request.get_json()
    albums = data.get('albums', [])

    if not albums:
        return jsonify({'success': False, 'error': '没有数据可下载'}), 400

    # 创建工作簿
    wb = Workbook()

    # 删除默认 Sheet
    wb.remove(wb.active)

    # 创建汇总 Sheet
    summary_ws = wb.create_sheet('专辑汇总', 0)
    summary_ws.append(['序号', '专辑名称', '艺人', '发布日期', '歌曲数量', '链接'])

    # 设置汇总表头样式
    header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, album in enumerate(albums, 1):
        summary_ws.append([
            idx,
            album.get('album_name', ''),
            album.get('artist', ''),
            album.get('release_date', ''),
            album.get('track_count', 0),
            album.get('source_url', '')
        ])

    # 调整汇总列宽
    summary_ws.column_dimensions['A'].width = 8
    summary_ws.column_dimensions['B'].width = 40
    summary_ws.column_dimensions['C'].width = 20
    summary_ws.column_dimensions['D'].width = 15
    summary_ws.column_dimensions['E'].width = 12
    summary_ws.column_dimensions['F'].width = 60

    # 为每个专辑创建详细 Sheet
    for album in albums:
        # 清理 Sheet 名称（Excel 限制 31 字符，不能包含特殊字符）
        sheet_name = album.get('album_name', 'Unknown')[:20]
        sheet_name = sheet_name.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '')

        # 确保名称唯一
        original_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name[:17]}_{counter}"
            counter += 1

        ws = wb.create_sheet(sheet_name)

        # 专辑信息
        ws.append(['专辑名称', album.get('album_name', '')])
        ws.append(['艺人', album.get('artist', '')])
        ws.append(['发布日期', album.get('release_date', '')])
        ws.append(['歌曲数量', album.get('track_count', 0)])
        ws.append([])  # 空行

        # 歌曲列表表头
        ws.append(['序号', '歌曲名称', '时长', '播放链接', '预览音频'])
        header_fill = PatternFill(start_color='007AFF', end_color='007AFF', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 歌曲数据
        for track in album.get('tracks', []):
            ws.append([
                track.get('index', ''),
                track.get('name', ''),
                track.get('duration', ''),
                track.get('play_url', ''),
                track.get('preview_url', '')
            ])

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50

        # 设置对齐
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
            row[0].alignment = Alignment(horizontal='center')  # 序号居中
            row[2].alignment = Alignment(horizontal='center')  # 时长居中

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"apple_music_albums_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.errorhandler(404)
def not_found(error):
    return jsonify({'success': False, 'error': '接口不存在'}), 404


@app.errorhandler(500)
def internal_error(error):
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


if __name__ == '__main__':
    # 确保模板目录存在
    os.makedirs('templates', exist_ok=True)
    os.makedirs('static/css', exist_ok=True)
    os.makedirs('static/js', exist_ok=True)

    app.run(debug=True, host='0.0.0.0', port=8765)
